import os
import requests
import pandas as pd
from requests.auth import HTTPBasicAuth
import xml.etree.ElementTree as ET
import json
import openpyxl
from datetime import datetime

def EcosysPOHeadersAPIData( api_url, username,password):
    # Create directory if it doesn't exist
    if not os.path.exists('Ecosys Data'):
        os.makedirs('Ecosys Data')

    # Set API URL and credentials
    #api_url = 'https://ecosys-stg.sbmoffshore.com/ecosys/api/restjson/EcosysPOLinesData_DCTAPI_KOL/?RootCostObject={RootCostObject}'
    #username = 'your_username'
    #password = 'your_password'

    # Set headers for JSON format
    headers = {'Content-type': 'application/json', 'Accept': 'application/json'}

    # Make API request with basic authentication
    response = requests.get(api_url, headers=headers, auth=HTTPBasicAuth(username, password))

    # Convert response from JSON to Python dictionary
    data = json.loads(response.content)

    # Get XML file from API response
    xml_string = data['xml']

    # Parse XML string into an ElementTree object
    root = ET.fromstring(xml_string)

    # Create Excel workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Get the current timestamp
    timestamp = datetime.now().strftime("%Y-%m-%d %H-%M-%S")

    # Write XML data to Excel sheet
    for child in root:
        ws.append([child.tag, child.attrib])

    # Save Excel workbook to file
    wb.save(f'Ecosys Data/Ecosys_POHeader{timestamp}.xlsx')

#--------------------------------------------------------------------------------------------------------------------
#--------------------------------------------------------------------------------------------------------------------

def EcosysPOLineAPIData(api_url,username,password):
    # Create directory if it doesn't exist
    if not os.path.exists('Ecosys Data'):
        os.makedirs('Ecosys Data')

    # Set API URL and credentials
    #api_url = 'https://example.com/api'
    #username = 'your_username'
    #password = 'your_password'

    # Set headers for JSON format
    headers = {'Content-type': 'application/json', 'Accept': 'application/json'}

    # Make API request with basic authentication
    response = requests.get(api_url, headers=headers, auth=HTTPBasicAuth(username, password))

    # Convert response from JSON to Python dictionary
    data = json.loads(response.content)

    # Get XML file from API response
    xml_string = data['xml']

    # Parse XML string into an ElementTree object
    root = ET.fromstring(xml_string)

    # Create Excel workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Get the current timestamp
    timestamp = datetime.now().strftime("%Y-%m-%d %H-%M-%S")

    # Write XML data to Excel sheet
    for child in root:
        ws.append([child.tag, child.attrib])

    # Save Excel workbook to file
    wb.save(f'Ecosys Data/Ecosys_POLine{timestamp}.xlsx')

    def EcosysPOLineAPIData(api_url, username, password):
        # Create directory if it doesn't exist
        if not os.path.exists('Ecosys Data'):
            os.makedirs('Ecosys Data')

        # Set API URL and credentials
        # api_url = 'https://example.com/api'
        # username = 'your_username'
        # password = 'your_password'

        # Set headers for JSON format
        headers = {'Content-type': 'application/json', 'Accept': 'application/json'}

        # Make API request with basic authentication
        response = requests.get(api_url, headers=headers, auth=HTTPBasicAuth(username, password))

        # Convert response from JSON to Python dictionary
        data = json.loads(response.content)

        # Get XML file from API response
        xml_string = data['xml']

        # Parse XML string into an ElementTree object
        root = ET.fromstring(xml_string)

        # Create Excel workbook and sheet
        wb = openpyxl.Workbook()
        ws = wb.active

        # Get the current timestamp
        timestamp = datetime.now().strftime("%Y-%m-%d %H-%M-%S")

        # Write XML data to Excel sheet
        for child in root:
            ws.append([child.tag, child.attrib])

        # Save Excel workbook to file
        wb.save(f'Ecosys Data/Ecosys_POLine{timestamp}.xlsx')

#--------------------------------------------------------------------------------------------------------------------
#--------------------------------------------------------------------------------------------------------------------

def EcosysSUNAPIData(API, username, password):
    try:
        # Make API request and get JSON response
        response = requests.get(API, auth=(username, password), verify=False)
        response.raise_for_status()
        json_data = json.loads(response.content)

        # Convert JSON data to Pandas DataFrame and select desired columns
        df = pd.DataFrame(json_data['WorkingForecastTransactionList'])

        # select desired columns
        df = df[['CostObjectHierarchyPathID', 'ExternalKey', 'TransactionReference', 'Description',
                 'InvoiceReference', 'PONumber', 'AccountCodeID', 'AccountCodeName', 'BusinessUnitID',
                 'CostCodeID', 'JournalTypeID', 'JournalNumber', 'JournalLine', 'JournalSourceID',
                 'CompanyID', 'SunSystemsPeriodID', 'TransactionDate', 'SunSystemsTransactionDate',
                 'Amount', 'Currency', 'ProjectAmount', 'ProjectCurrencyID', 'EntryDate',
                 'InvoiceHyperlink', 'SunSystemsPostingDate', 'SunSystemsBaseAmount', 'SunSystemsCpyCurrency']]

        # Get the current timestamp
        timestamp = datetime.now().strftime("%Y-%m-%d %H-%M-%S")

        # create an Excel writer object
        writer = pd.ExcelWriter(f'C:\\Users\\keven.deoliveiralope\\Documents\\Data Analyze Automatization\\Scripts\\Data-Collection-Transformation-kevLopes-DCT\\Automatized Process\\Data Pool\\Ecosys API Data\\SUN Transactions\\Ecosys SUN Transactions_{timestamp}.xlsx')

        # write the DataFrame to the Excel file
        df.to_excel(writer, index=False)

        # save the Excel file
        writer.save()
        print('Loading Data . . .')
        print('Data saved to Excel file successfully.')

    except requests.exceptions.RequestException as re:
        print('RequestException:', re)
    except Exception as e:
        print('Error:', e)


#--------------------------------------------------------------------------------------------------------------------
#--------------------------------------------------------------------------------------------------------------------

def ecosys_po_lines_data_api(api, username, password):
    try:
        # Make API request and get JSON response
        response = requests.get(api, auth=(username, password), verify=False)
        response.raise_for_status()
        json_data = json.loads(response.content)

        # Convert JSON data to Pandas DataFrame and select desired columns
        df = pd.DataFrame(json_data['WorkingForecastTransactionList'])
        df = df[['CostObjectHierarchyPathID', 'CostObjectID', 'CostBreakdownStructureHierarchyPathID',
                 'PORevisionID', 'TransactionDate', 'POLineNumber', 'PODescription', 'UnitofMeasureID',
                 'CostCostObjectCurrency', 'CurrencyCostObjectCode', 'CostTransactionCurrency',
                 'CurrencyTransactionCode', 'TagNumber', 'CostElementROSDate']]

        # Get the current timestamp
        timestamp = datetime.now().strftime("%Y-%m-%d %H-%M-%S")

        # Save DataFrame to Excel file
        writer = pd.ExcelWriter(f'C:\\Users\\keven.deoliveiralope\\Documents\\Data Analyze Automatization\\Scripts\\Data-Collection-Transformation-kevLopes-DCT\\Automatized Process\\Data Pool\\Ecosys API Data\\PO Lines\\Ecosys PO Lines Data_{timestamp}.xlsx')
        df.to_excel(writer, index=False)
        writer.save()

        print('Loading Data . . .')
        print("Excel file saved successfully!")

    except requests.exceptions.RequestException as e:
        print("Request error:", e)
    except (KeyError, ValueError) as e:
        print("Error parsing JSON data:", e)
    except Exception as e:
        print("An error occurred:", e)