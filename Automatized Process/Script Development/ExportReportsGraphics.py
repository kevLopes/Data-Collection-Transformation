import os
import openpyxl
from openpyxl.chart import ScatterChart, Reference, Series
from datetime import datetime
import matplotlib.pyplot as plt
import seaborn as sns

def data_graphic_reports_func():
    # Ask the user for the project ID
    project_id = input("Please enter the Project ID: ")

    # Find the Materials Data Organized folder
    folder_path = os.path.join(os.getcwd(), "Materials Data Organized")
    if not os.path.exists(folder_path):
        print(f"Error: Folder '{folder_path}' not found.")
        exit()

    # Find all Excel files with the project ID in the name
    files = [f for f in os.listdir(folder_path) if project_id in f and f.endswith('.xlsx')]

    # Define the columns to analyze
    columns = ["Tag Number", "ID", "Project Number", "Product Code", "Commodity Code",
               "Service Description", "Pipe Base Material", "Material", "LineNumber",
               "SBM scope", "Total QTY to commit", "Quantity UOM", "Unit Weight",
               "Unit Weight UOM", "Total NET weight", "SIZE"]

    # Create a dictionary to store the data
    data_dict = {}
    for column in columns:
        data_dict[column] = []

    # Loop through each file and extract the data
    for file in files:
        # Open the file
        try:
            workbook = openpyxl.load_workbook(os.path.join(folder_path, file))
        except Exception as e:
            print(f"Error: Could not load workbook {file} - {e}")
            continue

        # Loop through each worksheet
        for worksheet in workbook:
            # Loop through each row in the worksheet
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                # Loop through each column in the row
                for i, column in enumerate(columns):
                    data_dict[column].append(row[i])

    # Create a new workbook to store the results
    result_workbook = openpyxl.Workbook()

    # Create a worksheet to store the data
    data_worksheet = result_workbook.create_sheet(title="Data")

    # Write the data to the worksheet
    for i, column in enumerate(columns):
        data_worksheet.cell(row=1, column=i + 1, value=column)
        for j, value in enumerate(data_dict[column]):
            data_worksheet.cell(row=j + 2, column=i + 1, value=value)

    # Create a scatter chart for each file
    chart = ScatterChart()
    chart.title = f"{project_id} - Total NET weight vs Unit Weight"
    chart.x_axis.title = 'Total NET weight'
    chart.y_axis.title = 'Unit Weight'
    chart.legend = None

    for file in files:
        # Create a new series for the file
        series = Series(
            Reference(data_worksheet, min_col=columns.index("Total NET weight") + 1, min_row=2,
                      max_row=len(data_dict[columns[0]])),
            Reference(data_worksheet, min_col=columns.index("Unit Weight") + 1, min_row=2,
                      max_row=len(data_dict[columns[0]])),
            title=file
        )

        # Add the series to the chart
        chart.series.append(series)

    # Add the chart to the worksheet
    chart_worksheet = result_workbook.create_sheet(title="Chart")
    chart_worksheet.add_chart(chart, "A1")

    # Get the current timestamp
    timestamp = datetime.now().strftime("%Y-%m-%d %H-%M-%S")

    # Save the result workbook
    result_filename = f"Result Graphic {project_id}_{timestamp}.xlsx"
    result_workbook.save(os.path.join(folder_path, result_filename))

    print("Done!")


def plot_totals(cost_df, project_number):
    commit_totals = cost_df.groupby('Quantity UOM')['Total QTY to commit'].sum()
    po_totals = cost_df.groupby('UOM in PO')['Quantity in PO'].sum()
    currency_totals = cost_df.groupby('Transaction Currency')['PO Cost'].sum()

    # Create graphics directory if not exists
    graphics_dir = "../Data Pool/DCT Process Results/graphics"
    os.makedirs(graphics_dir, exist_ok=True)

    fig, axs = plt.subplots(3, 1, figsize=(12, 15))

    sns.barplot(x=commit_totals.index, y=commit_totals.values, ax=axs[0])
    axs[0].set_title('Total Committed Quantity per MTO')
    axs[0].set_xlabel('Quantity UOM')
    axs[0].set_ylabel('Total Committed Quantity')

    # Add value labels
    for p in axs[0].patches:
        axs[0].text(p.get_x() + p.get_width() / 2., p.get_height(),
               '%d' % int(p.get_height()),
               fontsize=12, color='black', ha='center', va='bottom')

    sns.barplot(x=po_totals.index, y=po_totals.values, ax=axs[1])
    axs[1].set_title('Total Quantity from Purchase Order')
    axs[1].set_xlabel('UOM in PO')
    axs[1].set_ylabel('Total Quantity in PO')

    # Add value labels
    for p in axs[1].patches:
        axs[1].text(p.get_x() + p.get_width() / 2., p.get_height(),
               '%d' % int(p.get_height()),
               fontsize=12, color='black', ha='center', va='bottom')

    sns.barplot(x=currency_totals.index, y=currency_totals.values, ax=axs[2])
    axs[2].set_title('Total Cost per Transaction Currency')
    axs[2].set_xlabel('Transaction Currency')
    axs[2].set_ylabel('Total Cost')

    # Add value labels
    for p in axs[2].patches:
        axs[2].text(p.get_x() + p.get_width() / 2., p.get_height(),
               '%d' % int(p.get_height()),
               fontsize=12, color='black', ha='center', va='bottom')

    plt.tight_layout()

    # Save figure
    timestamp = datetime.now().strftime("%Y-%m-%d %H-%M-%S")
    fig_name = f"MP{project_number}_PippingAnalyze_Graphics_Illustration_{timestamp}.png"
    fig_path = os.path.join(graphics_dir, fig_name)
    plt.savefig(fig_path)

    plt.tight_layout()
    plt.show()

    print(f'Figure saved at {fig_path}')

def plot_totals_other(cost_df, project_number):
    commit_totals = cost_df.groupby('Quantity UOM')['Total QTY to commit'].sum()
    po_totals = cost_df.groupby('UOM in PO')['Quantity in PO'].sum()
    currency_totals = cost_df.groupby('Transaction Currency')['PO Cost'].sum()

    # New: calculate cost per weight unit and group by Transaction Currency
    cost_df['Cost Per Weight'] = cost_df['Project Currency Cost'] / (
                cost_df['Total NET weight'] + cost_df['Total Weight using PO quantity'])
    cost_per_weight_totals = cost_df.groupby('Transaction Currency')['Cost Per Weight'].sum()

    # Create graphics directory if not exists
    graphics_dir = "../Data Pool/DCT Process Results/Graphics"
    os.makedirs(graphics_dir, exist_ok=True)

    fig, axs = plt.subplots(4, 1, figsize=(12, 20))  # Updated: 4 subplots instead of 3

    sns.barplot(x=commit_totals.index, y=commit_totals.values, ax=axs[0])
    axs[0].set_title('Total Committed Quantity per Quantity UOM')
    axs[0].set_xlabel('Quantity UOM')
    axs[0].set_ylabel('Total Committed Quantity')
    # Add value labels
    for p in axs[0].patches:
        axs[0].text(p.get_x() + p.get_width() / 2., p.get_height(),
                    '%d' % int(p.get_height()),
                    fontsize=12, color='black', ha='center', va='bottom')

    sns.barplot(x=po_totals.index, y=po_totals.values, ax=axs[1])
    axs[1].set_title('Total Quantity in PO per UOM in PO')
    axs[1].set_xlabel('UOM in PO')
    axs[1].set_ylabel('Total Quantity in PO')
    # Add value labels
    for p in axs[1].patches:
        axs[1].text(p.get_x() + p.get_width() / 2., p.get_height(),
                    '%d' % int(p.get_height()),
                    fontsize=12, color='black', ha='center', va='bottom')

    sns.barplot(x=currency_totals.index, y=currency_totals.values, ax=axs[2])
    axs[2].set_title('Total Cost per Transaction Currency')
    axs[2].set_xlabel('Transaction Currency')
    axs[2].set_ylabel('Total Cost')
    # Add value labels
    for p in axs[2].patches:
        axs[2].text(p.get_x() + p.get_width() / 2., p.get_height(),
                    '%d' % int(p.get_height()),
                    fontsize=12, color='black', ha='center', va='bottom')

    # New: plot cost per weight unit
    sns.barplot(x=cost_per_weight_totals.index, y=cost_per_weight_totals.values, ax=axs[3])
    axs[3].set_title('Cost per Weight Unit per Transaction Currency')
    axs[3].set_xlabel('Transaction Currency')
    axs[3].set_ylabel('Cost per Weight Unit')

    # Add value labels
    for p in axs[3].patches:
        axs[3].text(p.get_x() + p.get_width() / 2., p.get_height(),
                    '%.2f' % float(p.get_height()),
                    fontsize=12, color='black', ha='center', va='bottom')

    plt.tight_layout()

    # Save figure
    timestamp = datetime.now().strftime("%Y-%m-%d %H-%M-%S")
    fig_name = f"MP{project_number}_PippingAnalyze_Graphics_Illustration_{timestamp}.png"
    fig_path = os.path.join(graphics_dir, fig_name)
    plt.savefig(fig_path)

    plt.tight_layout()
    plt.show()

    print(f'Figure saved at {fig_path}')


def plot_cost_per_weight(cost_df, project_number):
    # Calculate cost per weight for each row
    cost_df['Cost per Weight'] = cost_df['Project Currency Cost'] / cost_df['Total Weight using PO quantity']

    # Aggregate totals
    cost_totals = cost_df['Project Currency Cost'].sum()
    net_weight_totals = cost_df['Total NET weight'].sum()
    weight_totals = cost_df['Total Weight using PO quantity'].sum()

    # Calculate total cost per weight
    total_cost_per_weight = cost_totals / weight_totals

    # Create graphics directory if not exists
    graphics_dir = "../Data Pool/DCT Process Results/graphics"
    os.makedirs(graphics_dir, exist_ok=True)

    # Create figure
    fig, ax = plt.subplots(figsize=(12, 5))

    # Plot total cost per weight
    sns.barplot(x=['Cost per Weight'], y=[total_cost_per_weight], ax=ax)

    ax.set_title('Total Cost per Weight')
    ax.set_ylabel('Cost per Weight')

    # Add value labels
    for p in ax.patches:
        ax.text(p.get_x() + p.get_width() / 2., p.get_height(),
               '%.2f' % float(p.get_height()),
               fontsize=12, color='black', ha='center', va='bottom')

    plt.tight_layout()

    # Save figure
    fig_name = f"MP{project_number}_Total_Cost_Per_Weight.png"
    fig_path = os.path.join(graphics_dir, fig_name)
    plt.savefig(fig_path)
    print(f'Figure saved at {fig_path}')
