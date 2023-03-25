import os
import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.label import DataLabelList

# Ask the user for the project ID
project_id = input("Please enter the Project ID: ")

# Find the Materials Data Organized folder
folder_path = os.path.join(os.getcwd(), "Materials Data Organized")
if not os.path.exists(folder_path):
    print(f"Error: Folder '{folder_path}' not found.")
    exit()

# Find all Excel files with the project ID in the name
files = [f for f in os.listdir(folder_path) if project_id in f and f.endswith('.xlsx')]

# Create a new workbook to store the results
result_workbook = openpyxl.Workbook()

# Create the first worksheet for the first graph
worksheet1 = result_workbook.active
worksheet1.title = "Graph 1"

# Create the second worksheet for the second graph
worksheet2 = result_workbook.create_sheet(title="Graph 2")

# Loop through each file and extract the data
for file in files:
    # Open the file
    workbook = openpyxl.load_workbook(os.path.join(folder_path, file))

    # Extract the data from the first worksheet
    worksheet = workbook.active
    data = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        data.append(row)

    # Create the first graph
    chart1 = LineChart()
    chart1.title = f"{file} - Graph 1"
    chart1.x_axis.title = 'X Axis'
    chart1.y_axis.title = 'Y Axis'
    chart1.data_labels = DataLabelList()
    chart1.data_labels.show_val = True

    x_data = Reference(worksheet, min_col=1, min_row=2, max_row=len(data))
    y_data = Reference(worksheet, min_col=2, min_row=2, max_row=len(data))
    chart1.add_data(y_data, titles_from_data=True)
    chart1.set_categories(x_data)

    # Add the first graph to the first worksheet
    worksheet1.add_chart(chart1, f"A{worksheet1.max_row+2}")

    # Create the second graph
    chart2 = LineChart()
    chart2.title = f"{file} - Graph 2"
    chart2.x_axis.title = 'X Axis'
    chart2.y_axis.title = 'Y Axis'
    chart2.data_labels = DataLabelList()
    chart2.data_labels.show_val = True

    x_data = Reference(worksheet, min_col=1, min_row=2, max_row=len(data))
    y_data = Reference(worksheet, min_col=3, min_row=2, max_row=len(data))
    chart2.add_data(y_data, titles_from_data=True)
    chart2.set_categories(x_data)

    # Add the second graph to the second worksheet
    worksheet2.add_chart(chart2, f"A{worksheet2.max_row+2}")

# Save the result workbook
result_filename = f"Result Graphic {project_id}.xlsx"
result_workbook.save(os.path.join(folder_path, result_filename))

print("Done!")
