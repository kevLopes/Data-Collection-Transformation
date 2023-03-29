import os
from pathlib import Path
import openpyxl


def data_analyze_process():
    # Ask for project ID input
    project_id = input("Please enter the Project ID: ")

    # Find and open Excel file in Data Pool folder
    folder_path = os.path.join(os.getcwd(),
                               "/Users/lopes/PycharmProjects/PythonAutomatization/Automatized Process/Data Pool/Materials Data Organized")
    file_name = f"{project_id} Piping.xlsx"
    file_path = os.path.join(folder_path, file_name)

    if os.path.exists(file_path):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # Find the Tag Number column
        tag_number_col = None
        for col in range(1, sheet.max_column + 1):
            col_name = sheet.cell(row=1, column=col).value
            if col_name == "Tag Number":
                tag_number_col = col
                break

        if tag_number_col is None:
            print("Tag Number column not found in Piping file")
        else:
            # Get list of all Tag Number values
            tag_numbers = []
            for row in range(2, sheet.max_row + 1):
                tag_id_number = sheet.cell(row=row, column=tag_number_col).value
                if tag_id_number is not None:
                    tag_numbers.append(tag_id_number)

            # Get list of all Tag Number values
            for row in range(2, sheet.max_row + 1):
                tag_id_number = sheet.cell(row=row, column=1).value
                if tag_id_number is None:
                    continue

                # Open new excel file containing Project ID and PO Lines
                po_folder_path = Path(
                    os.getcwd()) / "Users" / "lopes" / "PycharmProjects" / "PythonAutomatization" / "Automatized Process" / "Ecosys Data"
                po_file_path = po_folder_path / f"{project_id} PO Lines.xlsx"

                # po_folder_path = os.path.join(os.getcwd(), "/Users/lopes/PycharmProjects/PythonAutomatization/Automatized Process/Ecosys Data")
                # po_file_name = f"{project_id} PO Lines.xlsx"
                # po_file_path = os.path.join(po_folder_path, po_file_name)

                if os.path.exists(po_file_path):
                    po_workbook = openpyxl.load_workbook(po_file_path)
                    po_sheet = po_workbook.active

                    # Search for Tag Number column in PO Lines file
                    tag_number_col = None
                    for col in range(1, po_sheet.max_column + 1):
                        col_name = po_sheet.cell(row=1, column=col).value
                        if col_name == "Tag Number":
                            tag_number_col = col
                            break

                    if tag_number_col is None:
                        print("Tag Number column not found in PO Lines file")
                        continue

                    # Find matching row in PO Lines file
                    for po_row in range(2, po_sheet.max_row + 1):
                        po_tag_number = po_sheet.cell(row=po_row, column=tag_number_col).value
                        if po_tag_number == tag_id_number:

                            # Check if file already exists
                            new_file_name = f"{project_id} Data Analysed.xlsx"
                            new_file_path = os.path.join(folder_path, new_file_name)
                            new_workbook = None
                            new_sheet = None

                            if os.path.exists(new_file_path):
                                new_workbook = openpyxl.load_workbook(new_file_path)
                                new_sheet = new_workbook.active
                            else:
                                new_workbook = openpyxl.Workbook()
                                new_sheet = new_workbook.active

                                # Add headers to new file
                                for col in range(1, sheet.max_column + po_sheet.max_column + 1):
                                    if col <= sheet.max_column:
                                        header = sheet.cell(row=1, column=col).value
                                    else:
                                        header = po_sheet.cell(row=1, column=col - sheet.max_column).value
                                    new_sheet.cell(row=1, column=col, value=header)

                            # Add data to new file
                            for col in range(1, sheet.max_column + po_sheet.max_column + 1):
                                if col <= sheet.max_column:
                                    value = sheet.cell(row=row, column=col).value
                                else:
                                    value = po_sheet.cell(row=po_row, column=col - sheet.max_column).value
                                new_sheet.cell(row=new_sheet.max_row + 1, column=col, value=value)

                            new_workbook.save(new_file_path)

                else:
                    print(f"PO Lines file not found for {project_id}")
    else:
        print(f"Piping file not found for {project_id}")
