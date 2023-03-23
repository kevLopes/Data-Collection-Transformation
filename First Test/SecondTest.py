import os
from datetime import datetime
import openpyxl

# Load the source workbook
source_wb = openpyxl.load_workbook("/Users/lopes/PycharmProjects/PythonAutomatization/First Test/Fichier 1.xlsx")

# Select the active worksheet in the source workbook
source_ws = source_wb.active

# Get the value in cell A5 of the source worksheet
value_to_compare = source_ws["A5"].value

# Load the destination workbook
dest_wb = openpyxl.load_workbook("/Users/lopes/PycharmProjects/PythonAutomatization/First Test/Fichier 2.xlsx")

# Select the active worksheet in the destination workbook
dest_ws = dest_wb.active

# Get the value in cell B8 of the destination worksheet
value_in_dest = dest_ws["B8"].value

# Compare the values
if value_to_compare == value_in_dest:
    # Create a new workbook
    result_wb = openpyxl.Workbook()

    # Select the active worksheet in the result workbook
    result_ws = result_wb.active

    # Write the value in cell C3 of the result worksheet
    result_ws["C3"].value = value_to_compare

    # Get the current timestamp
    timestamp = datetime.now().strftime("%Y-%m-%d %H-%M-%S")

    # Create the directory if it doesn't exist
    result_dir = "/Users/lopes/PycharmProjects/PythonAutomatization/First Test/Script 1"
    if not os.path.exists(result_dir):
        os.makedirs(result_dir)

    # Save the result workbook with timestamp in filename
    result_wb.save(os.path.join(result_dir, f"Result_{timestamp}.xlsx"))
else:
    # Create a new workbook
    not_match_wb = openpyxl.Workbook()

    # Select the active worksheet in the not match workbook
    not_match_ws = not_match_wb.active

    # Write the value in cell D9 of the not match worksheet
    not_match_ws["D9"].value = value_to_compare

    # Get the current timestamp
    timestamp = datetime.now().strftime("%Y-%m-%d %H-%M-%S")

    # Create the directory if it doesn't exist
    result_dir = "/Users/lopes/PycharmProjects/PythonAutomatization/First Test/Script 1"
    if not os.path.exists(result_dir):
        os.makedirs(result_dir)

    # Save the result workbook with timestamp in filename
    not_match_wb.save(os.path.join(result_dir, f"NOT Match{timestamp}.xlsx"))