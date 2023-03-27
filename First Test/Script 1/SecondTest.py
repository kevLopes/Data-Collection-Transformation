import openpyxl

# Open the first workbook and get the value from cell A5
wb1 = openpyxl.load_workbook(r'C:\Users\keven.deoliveiralope\Documents\Data Analyze Automatization\Scripts\Script 1\Fichie 1.xlsx')
ws1 = wb1.active
value_to_compare = ws1['A5'].value

# Open the second workbook and get the value from cell B8
wb2 = openpyxl.load_workbook(r'C:\Users\keven.deoliveiralope\Documents\Data Analyze Automatization\Scripts\Script 1\Fichie 2.xlsx')
ws2 = wb2.active
value_in_B8 = ws2['B8'].value

# Compare the values
if value_to_compare == value_in_B8:
    # Create a new workbook and put the value in cell C3
    wb3 = openpyxl.Workbook()
    ws3 = wb3.active
    ws3['C3'].value = value_to_compare
    wb3.save(r'C:\Users\keven.deoliveiralope\Documents\Data Analyze Automatization\Scripts\Script 1\Result.xlsx')
else:
    # Create a new workbook and put the value in cell D9
    wb3 = openpyxl.Workbook()
    ws3 = wb3.active
    ws3['D9'].value = value_to_compare
    wb3.save(r'C:\Users\keven.deoliveiralope\Documents\Data Analyze Automatization\Scripts\Script 1\NOT Match.xlsx')
